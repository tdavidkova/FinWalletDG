"""
Excel backup utility — exports a full snapshot of all data every time
a mutating operation occurs.  Files are written to the ``backups/`` folder
with a timestamp in the filename.
"""

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.models import Group, Student, Transaction, ExpenseCategory, Invoice

BACKUP_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "backups")


def _ensure_dir():
    os.makedirs(BACKUP_DIR, exist_ok=True)


HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def _student_balance(db: Session, student_id: str):
    row = (
        db.query(
            func.coalesce(func.sum(Transaction.amount_bgn), 0),
            func.coalesce(func.sum(Transaction.amount_eur), 0),
        )
        .filter(Transaction.student_id == student_id)
        .first()
    )
    return round(row[0], 2), round(row[1], 2)


def save_backup(db: Session):
    """Export every table to a multi-sheet Excel file in ``backups/``."""
    _ensure_dir()

    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S_%f")
    path = os.path.join(BACKUP_DIR, f"backup_{ts}.xlsx")

    wb = Workbook()

    # ---- Sheet 1: Groups ----
    ws = wb.active
    ws.title = "Групи"
    headers = ["ID", "Име", "Детска градина", "Обменен курс", "Създадена"]
    ws.append(headers)
    for g in db.query(Group).order_by(Group.name).all():
        ws.append([g.id, g.name, g.kindergarten_name, g.exchange_rate, str(g.created_at)])
    _style_header(ws, len(headers))
    _auto_width(ws)

    # ---- Sheet 2: Students ----
    ws = wb.create_sheet("Деца")
    headers = ["ID", "Група", "Име", "Номер", "Статус", "Отписан на", "Sibling Group", "Баланс BGN", "Баланс EUR"]
    ws.append(headers)
    for s in db.query(Student).order_by(Student.group_id, Student.display_number).all():
        bal_bgn, bal_eur = _student_balance(db, s.id)
        ws.append([
            s.id, s.group_id, s.full_name, s.display_number,
            s.status, str(s.unenrolled_at) if s.unenrolled_at else "",
            s.sibling_group_id or "", bal_bgn, bal_eur,
        ])
    _style_header(ws, len(headers))
    _auto_width(ws)

    # ---- Sheet 3: Transactions ----
    ws = wb.create_sheet("Движения")
    headers = ["ID", "Дете ID", "Дете", "Сума BGN", "Сума EUR", "Дата", "Основание", "Категория", "Batch ID", "Създадена"]
    ws.append(headers)
    for tx in db.query(Transaction).order_by(Transaction.date.desc(), Transaction.created_at.desc()).all():
        cat_name = tx.category.name if tx.category else ""
        ws.append([
            tx.id, tx.student_id, tx.student.full_name,
            tx.amount_bgn, tx.amount_eur,
            str(tx.date), tx.reason, cat_name,
            tx.expense_batch_id or "", str(tx.created_at),
        ])
    _style_header(ws, len(headers))
    _auto_width(ws)

    # ---- Sheet 4: Categories ----
    ws = wb.create_sheet("Категории")
    headers = ["ID", "Група ID", "Име"]
    ws.append(headers)
    for c in db.query(ExpenseCategory).order_by(ExpenseCategory.group_id, ExpenseCategory.name).all():
        ws.append([c.id, c.group_id, c.name])
    _style_header(ws, len(headers))
    _auto_width(ws)

    # ---- Sheet 5: Invoices ----
    ws = wb.create_sheet("Фактури")
    headers = ["ID", "Група ID", "Batch ID", "Обща сума", "На дете", "Описание", "Дата", "Номер фактура", "Валута", "Брой деца"]
    ws.append(headers)
    for inv in db.query(Invoice).order_by(Invoice.date.desc()).all():
        ws.append([
            inv.id, inv.group_id, inv.expense_batch_id or "",
            inv.total_amount, inv.per_child_cost,
            inv.description, str(inv.date),
            inv.invoice_number or "", inv.currency,
            inv.num_children,
        ])
    _style_header(ws, len(headers))
    _auto_width(ws)

    wb.save(path)
    return path
